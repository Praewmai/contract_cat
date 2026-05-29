"""
contract_to_excel.py
====================
Core logic for generating dashboard-ready Excel files from parsed contract data.
Input:  parsed_data dict (from AI extraction)
Output: .xlsx file matching dashboard upload format

Column order (must match dashboard):
_id, hotel_id, room_id, status, created_date, edited_date, start_date, end_date,
refundable, abf, contract_type, cutoff_date, hotel_supplier, important_message,
min_nights_stay, min_advance_days, net_price, net_price_emerald, net_price_ruby,
net_price_topaz, promo_book_till, promo_code, promo_note, room_allotment,
all_inclusive, baby_cot, cancellation_policy, cancellation_policy_net,
early_check_in, child_policy, child_share_bed_abf, child_extra_bed_abf,
extra_bed_abf, extra_bed_no_abf, full_board, half_board, hotel_extra_fees,
room_name, hotel_transfer, late_check_out, meals_and_info, tags, action
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import math


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

HEADERS = [
    '_id', 'hotel_id', 'room_id', 'status', 'created_date', 'edited_date',
    'start_date', 'end_date', 'refundable', 'abf', 'contract_type',
    'cutoff_date', 'hotel_supplier', 'important_message', 'min_nights_stay',
    'min_advance_days', 'net_price', 'net_price_emerald', 'net_price_ruby',
    'net_price_topaz', 'promo_book_till', 'promo_code', 'promo_note',
    'room_allotment', 'all_inclusive', 'baby_cot', 'cancellation_policy',
    'cancellation_policy_net', 'early_check_in', 'child_policy',
    'child_share_bed_abf', 'child_extra_bed_abf', 'extra_bed_abf',
    'extra_bed_no_abf', 'full_board', 'half_board', 'hotel_extra_fees',
    'room_name', 'hotel_transfer', 'late_check_out', 'meals_and_info',
    'tags', 'action'
]

COL_WIDTHS = {
    '_id': 10, 'hotel_id': 12, 'room_id': 16, 'status': 8,
    'created_date': 20, 'edited_date': 20, 'start_date': 14, 'end_date': 14,
    'refundable': 10, 'abf': 12, 'contract_type': 16, 'cutoff_date': 14,
    'hotel_supplier': 36, 'important_message': 22, 'min_nights_stay': 16,
    'min_advance_days': 17, 'net_price': 14, 'net_price_emerald': 16,
    'net_price_ruby': 15, 'net_price_topaz': 15, 'promo_book_till': 22,
    'promo_code': 18, 'promo_note': 30, 'room_allotment': 14,
    'all_inclusive': 14, 'baby_cot': 10, 'cancellation_policy': 32,
    'cancellation_policy_net': 28, 'early_check_in': 14, 'child_policy': 32,
    'child_share_bed_abf': 20, 'child_extra_bed_abf': 20, 'extra_bed_abf': 15,
    'extra_bed_no_abf': 16, 'full_board': 12, 'half_board': 12,
    'hotel_extra_fees': 18, 'room_name': 32, 'hotel_transfer': 16,
    'late_check_out': 15, 'meals_and_info': 32, 'tags': 8, 'action': 10,
}


# ─────────────────────────────────────────────
# ROW BUILDERS
# ─────────────────────────────────────────────

def _base_row(parsed, room, period, net_price, contract_type,
              promo_code=None, min_advance_days=None, min_nights_stay=None,
              promo_note=None, promo_book_till=None):
    """Build a single dashboard row dict."""
    
    allotment = period.get('room_allotment')
    if isinstance(allotment, dict):
        allotment = allotment.get(room['room_id'])
        
    return {
        '_id':                    None,
        'hotel_id':               parsed['hotel_id'],
        'room_id':                room['room_id'],
        'status':                 True,                          # always TRUE
        'created_date':           datetime.now(),                # today's datetime
        'edited_date':            None,                          # left blank for user
        'start_date':             period['start_date'],
        'end_date':               period['end_date'],
        'refundable':             False,
        'abf':                    parsed.get('abf', 'Included'),
        'contract_type':          contract_type,
        'cutoff_date':            period.get('cutoff_date'),
        'hotel_supplier':         parsed['hotel_supplier'],
        'important_message':      None,
        'min_nights_stay':        min_nights_stay or period.get('min_nights_stay'),
        'min_advance_days':       min_advance_days,
        'net_price':              str(int(net_price)) if net_price is not None else '0',
        'net_price_emerald':      None,
        'net_price_ruby':         None,
        'net_price_topaz':        None,
        'promo_book_till':        promo_book_till,
        'promo_code':             promo_code,
        'promo_note':             promo_note,
        'all_inclusive':          None,
        'baby_cot':               None,
        'cancellation_policy':    parsed.get('cancellation_policy') or None,
        'cancellation_policy_net':None,
        'early_check_in':         None,
        'child_policy':           parsed.get('child_policy') or None,
        'child_share_bed_abf':    parsed.get('child_share_bed_abf'),
        'child_extra_bed_abf':    parsed.get('child_extra_bed_abf'),
        'extra_bed_abf':          parsed.get('extra_bed_abf'),
        'extra_bed_no_abf':       parsed.get('extra_bed_no_abf'),
        'full_board':             parsed.get('full_board'),
        'half_board':             parsed.get('half_board'),
        'hotel_extra_fees':       None,
        'room_name':              room['room_name'],
        'hotel_transfer':         None,
        'late_check_out':         None,
        'meals_and_info':         parsed.get('meals_and_info') or None,
        'room_allotment':         allotment,
        'tags':                   '[]',
        'action':                 'insert',
    }


def _surcharge_note(currency, amount, custom_note=None):
    """HTML bold surcharge note matching dashboard format."""
    if custom_note:
        return f'<p><strong>{custom_note}</strong></p>'
    return f'<p><strong>Room rate includes Surcharge {currency} {int(amount):,} per room per night</strong></p>'


def _weekday_note(is_weekend, weekend_days="Saturday", weekday_days="Sun-Fri"):
    """HTML weekday/weekend note matching dashboard format exactly."""
    if is_weekend:
        return f'<p><span style="color: #ff0000;"><strong>Rate for Weekend ({weekend_days})</strong></span></p>'
    return f'<p><span style="color: #ff0000;"><strong>Rate for Weekday ({weekday_days})</strong></span></p>'


# ─────────────────────────────────────────────
# CORE GENERATOR
# ─────────────────────────────────────────────

def generate_rows(parsed, selected_cts):
    """
    Generate all Excel rows from parsed contract data.

    parsed        : dict from AI extraction
    selected_cts  : list of dicts e.g.
                    [{'type':'main'}, {'type':'eb','code':'E.B DAYS'},
                     {'type':'promo','promo_code':'EBO','promo_till':'2026-09-30'},
                     {'type':'por'}]

    Returns list of row dicts matching HEADERS.
    """
    rows = []
    ct_types = {c['type'] for c in selected_cts}
    ct_map   = {c['type']: c for c in selected_cts}

    for period in parsed.get('periods', []):
        is_peak       = period.get('is_peak', False)
        is_blackout   = period.get('eb_blackout', False)
        has_surcharge = period.get('has_surcharge', False)
        surcharge_amt = period.get('surcharge_amount', 0) or 0
        surcharge_cur = period.get('surcharge_currency', '')
        # per-room surcharge dict (overrides surcharge_amount per room)
        surcharge_rates = period.get('surcharge_rates', {})
        has_ww        = period.get('has_weekday_weekend', False)
        weekend_days  = period.get('weekend_days', 'Saturday')
        weekday_days  = period.get('weekday_days', 'Sun-Fri')
        custom_surcharge_note = period.get('surcharge_note')
        eb_tiers      = period.get('early_bird_tiers', [])

        for room in parsed.get('rooms', []):
            rid = room['room_id']
            base_rate = (period.get('rates') or {}).get(rid, 0) or 0
            # Use per-room surcharge if available, else fall back to period-level
            room_surcharge = surcharge_rates.get(rid, surcharge_amt) if surcharge_rates else surcharge_amt

            # ── MAIN CONTRACT ──────────────────────────────
            if 'main' in ct_types:
                if has_ww:
                    # Weekday row (base rate = weekday)
                    rows.append(_base_row(
                        parsed, room, period, base_rate, 'Main Contract',
                        promo_code='Main - Weekday',
                        promo_note=_weekday_note(False, weekend_days, weekday_days)
                    ))
                    # Weekend row (base + per-room weekend surcharge)
                    weekend_rate = base_rate + room_surcharge if has_surcharge else base_rate
                    weekend_note = _weekday_note(True, weekend_days, weekday_days)
                    if has_surcharge and room_surcharge:
                        weekend_note += _surcharge_note(surcharge_cur, room_surcharge, custom_surcharge_note)
                    rows.append(_base_row(
                        parsed, room, period, weekend_rate, 'Main Contract',
                        promo_code='Main - Weekend',
                        promo_note=weekend_note
                    ))
                elif has_surcharge:
                    rows.append(_base_row(
                        parsed, room, period,
                        base_rate + room_surcharge, 'Main Contract',
                        promo_note=_surcharge_note(surcharge_cur, room_surcharge, custom_surcharge_note)
                    ))
                else:
                    rows.append(_base_row(
                        parsed, room, period, base_rate, 'Main Contract'
                    ))

            # ── EARLY BIRD ──────────────────────────────
            if 'eb' in ct_types and not is_peak and not is_blackout:
                for tier in eb_tiers:
                    days       = tier.get('days', 0)
                    discount   = tier.get('discount_pct', 0)
                    eb_rate    = round(base_rate * (1 - discount / 100))
                    # AI extracts promo_code from PDF; fallback = "E.B {days} DAYS"
                    promo_code = tier.get('promo_code') or f'E.B {days} DAYS'
                    # EB usually has no promo_book_till (unless PDF explicitly states one)
                    eb_book_till = tier.get('promo_book_till') or None

                    if has_surcharge:
                        eb_rate += room_surcharge
                        note = _surcharge_note(surcharge_cur, room_surcharge, custom_surcharge_note)
                    else:
                        note = None

                    rows.append(_base_row(
                        parsed, room, period, eb_rate, 'Early Bird',
                        promo_code=promo_code,
                        min_advance_days=days,
                        promo_note=note,
                        promo_book_till=eb_book_till
                    ))

            # ── PROMOTION ──────────────────────────────────
            if 'promo' in ct_types and not is_blackout:
                promo_cfg  = ct_map.get('promo', {})
                promotions = period.get('promotions', [])

                for promo in promotions:
                    discount   = promo.get('discount_pct', 0)
                    promo_rate = round(base_rate * (1 - discount / 100))
                    promo_code = promo.get('promo_code') or promo_cfg.get('promo_code', '')
                    book_till  = promo.get('promo_book_till') or promo_cfg.get('promo_till')

                    if has_ww:
                        # Weekday promo row
                        rows.append(_base_row(
                            parsed, room, period, promo_rate, 'Promotion',
                            promo_code=f'{promo_code} - Weekday',
                            promo_book_till=book_till,
                            promo_note=_weekday_note(False, weekend_days, weekday_days)
                        ))
                        # Weekend promo row (per-room surcharge)
                        weekend_promo = promo_rate + room_surcharge if has_surcharge else promo_rate
                        weekend_note  = _weekday_note(True, weekend_days, weekday_days)
                        if has_surcharge and room_surcharge:
                            weekend_note += _surcharge_note(surcharge_cur, room_surcharge, custom_surcharge_note)
                        rows.append(_base_row(
                            parsed, room, period, weekend_promo, 'Promotion',
                            promo_code=f'{promo_code} - Weekend',
                            promo_book_till=book_till,
                            promo_note=weekend_note
                        ))
                    else:
                        note = _surcharge_note(surcharge_cur, room_surcharge, custom_surcharge_note) if has_surcharge else None
                        rows.append(_base_row(
                            parsed, room, period, promo_rate, 'Promotion',
                            promo_code=promo_code,
                            promo_book_till=book_till,
                            promo_note=note
                        ))

            # ── POR ────────────────────────────────────────
            if 'por' in ct_types:
                rows.append(_base_row(
                    parsed, room, period, 0, 'POR',
                    promo_code='POR Rate'
                ))

    # Custom sort strictly matching the example format (by start_date, then room_name)
    rows.sort(key=lambda r: (r.get('start_date', ''), r.get('room_name', '')))
    return rows


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

def write_excel(rows, output_path):
    """Write rows to .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header row styling
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    data_font   = Font(name='Arial', size=10)

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(h, 12)

    ws.row_dimensions[1].height = 20

    # Data rows
    date_only_cols  = {'start_date', 'end_date', 'edited_date', 'cutoff_date'}
    datetime_cols   = {'created_date'}

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(HEADERS, 1):
            val  = row.get(h)
            if isinstance(val, dict):
                val = ", ".join([f"{k.capitalize()}: {v}" for k, v in val.items()])
            elif isinstance(val, list):
                val = ", ".join(map(str, val))
            cell = ws.cell(r_idx, c_idx, val)
            cell.font = data_font

            if h in date_only_cols and val:
                cell.number_format = 'DD/MM/YYYY'
            elif h in datetime_cols and val:
                cell.number_format = 'DD/MM/YYYY HH:MM:SS'

            if h == 'net_price' and val is not None:
                # Keep as string per dashboard requirement
                cell.value = str(val)

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────

def contract_to_excel(parsed_data, selected_cts, output_path):
    """
    Full pipeline: parsed data → Excel file.

    Usage:
        from contract_to_excel import contract_to_excel

        result = contract_to_excel(
            parsed_data   = ai_response,       # dict from AI
            selected_cts  = [
                {'type': 'main'},
                {'type': 'eb', 'code': 'E.B DAYS'},
                {'type': 'promo', 'promo_code': 'EBO', 'promo_till': '2026-09-30'},
                {'type': 'por'},
            ],
            output_path   = 'output.xlsx'
        )
        print(f"Created {result} with rows.")
    """
    rows = generate_rows(parsed_data, selected_cts)
    write_excel(rows, output_path)
    return output_path, len(rows)


# ─────────────────────────────────────────────
# EXAMPLE — run directly to test
# ─────────────────────────────────────────────

if __name__ == '__main__':
    # Minimal test data matching AI output format
    sample = {
        'hotel_id': '12711588',
        'hotel_supplier': 'Namia River Retreat - Wellness Inclusive Resort',
        'abf': 'Included',
        'extra_bed_abf': 4000000,
        'child_extra_bed_abf': 2500000,
        'half_board': 1620000,
        'full_board': 2700000,
        'cancellation_policy': '<p>Up to 20 days: No charge. Within 19 days: 100% charge.</p>',
        'child_policy': '<p>Child under 6: FOC. Child 6-15: Extra bed + ABF 2,500,000 VND.</p>',
        'meals_and_info': None,
        'rooms': [
            {'room_id': '1271158801', 'room_name': 'One Bedroom Nipa Pool Villa'},
            {'room_id': '1271158802', 'room_name': 'One Bedroom River Pool Villa'},
        ],
        'periods': [
            {
                'start_date': '2027-04-01',
                'end_date':   '2027-12-22',
                'season':     'Normal Season',
                'is_peak':    False,
                'eb_blackout': False,
                'has_surcharge': False,
                'surcharge_amount': 0,
                'surcharge_currency': '',
                'min_nights_stay': None,
                'has_weekday_weekend': False,
                'rates': {
                    '1271158801': 13200000,
                    '1271158802': 16830000,
                },
                'early_bird_tiers': [
                    {'days': 60, 'discount_pct': 15, 'promo_code': 'E.B 60 DAYS'},
                ],
                'promotions': [],
            },
            {
                'start_date': '2027-12-23',
                'end_date':   '2028-01-07',
                'season':     'Peak Dates',
                'is_peak':    True,
                'eb_blackout': True,
                'has_surcharge': False,
                'surcharge_amount': 0,
                'surcharge_currency': '',
                'min_nights_stay': None,
                'has_weekday_weekend': False,
                'rates': {
                    '1271158801': 15180000,
                    '1271158802': 19350000,
                },
                'early_bird_tiers': [],
                'promotions': [],
            },
        ],
    }

    selected = [
        {'type': 'main'},
        {'type': 'eb', 'code': 'E.B DAYS'},
        {'type': 'por'},
    ]

    path, count = contract_to_excel(sample, selected, '/home/claude/test_output.xlsx')
    print(f'✅ Created {path} — {count} rows')
