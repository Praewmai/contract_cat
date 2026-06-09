import streamlit as st
import json
import base64
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import os
from contract_to_excel import generate_rows, HEADERS, COL_WIDTHS

# ─── Page Config ───
st.set_page_config(
    page_title="MeowAI 🐾 Contract Extractor",
    page_icon="🐾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Mascot Setup (Transparent PNG) ───
try:
    from mascot_data import MASCOT_B64
except ImportError:
    MASCOT_B64 = ""

if MASCOT_B64:
    mascot_img_html = f'<div class="mascot-container"><img src="data:image/png;base64,{MASCOT_B64}" class="mascot-avatar" /></div>'
else:
    mascot_img_html = '<div class="mascot-container" style="display:flex;align-items:center;justify-content:center;font-size:80px;">🐱</div>'

# ─── CSS ───
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif !important;
    }

    /* ── Background ── */
    .stApp {
        background: radial-gradient(circle at top left, #F8FAFC 0%, #E2E8F0 100%) !important;
        min-height: 100vh;
    }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%) !important;
        border-right: 1px solid rgba(255,255,255,0.08) !important;
    }
    [data-testid="stSidebar"] * { color: #E2E8F0 !important; }
    [data-testid="stSidebar"] h4 { color: #94A3B8 !important; font-size: 0.8rem !important; text-transform: uppercase; letter-spacing: 1.5px; margin-top: 1rem !important; font-weight: 700; }
    [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.1) !important; margin: 1.5rem 0 !important; }
    [data-testid="stSidebar"] [data-baseweb="input"] {
        background-color: rgba(255, 255, 255, 0.08) !important;
        border-radius: 12px !important;
        border: 1px solid rgba(255, 255, 255, 0.15) !important;
        transition: all 0.3s ease;
    }
    [data-testid="stSidebar"] [data-baseweb="input"] > div {
        background-color: transparent !important;
    }
    [data-testid="stSidebar"] [data-baseweb="input"] input {
        background-color: transparent !important;
        color: #F8FAFC !important;
        border: none !important;
    }
    [data-testid="stSidebar"] [data-baseweb="input"]:focus-within {
        background-color: rgba(255, 255, 255, 0.12) !important;
        border-color: #94A3B8 !important;
        box-shadow: 0 0 0 3px rgba(255, 255, 255, 0.1) !important;
    }
    [data-testid="stSidebar"] [data-baseweb="input"] input::placeholder {
        color: rgba(255, 255, 255, 0.4) !important;
    }

    /* ── Mascot ── */
    .mascot-container {
        display: flex;
        justify-content: center;
        align-items: center;
        padding: 1.5rem 0 1rem 0;
    }
    .mascot-avatar {
        width: 140px;
        height: 140px;
        object-fit: contain;
        filter: drop-shadow(0 8px 20px rgba(0,0,0,0.3));
        animation: float-mascot 4s ease-in-out infinite;
        transition: transform 0.3s cubic-bezier(0.34, 1.56, 0.64, 1);
    }
    .mascot-avatar:hover {
        transform: scale(1.1) rotate(2deg);
    }
    @keyframes float-mascot {
        0%   { transform: translateY(0px); }
        50%  { transform: translateY(-8px); }
        100% { transform: translateY(0px); }
    }

    .sidebar-title {
        text-align: center;
        font-size: 1.6rem !important;
        font-weight: 800 !important;
        background: linear-gradient(90deg, #F8FAFC, #94A3B8);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0.2rem 0 0.1rem 0 !important;
    }
    .sidebar-subtitle {
        text-align: center;
        font-size: 0.8rem !important;
        color: #94A3B8 !important;
        margin: 0 !important;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        font-weight: 600;
    }
    .sidebar-pill {
        background: rgba(255,255,255,0.05);
        border-radius: 12px;
        padding: 0.6rem 1rem;
        text-align: center;
        font-size: 0.8rem;
        color: #E2E8F0 !important;
        margin: 1.2rem 0;
        border: 1px solid rgba(255,255,255,0.1);
        backdrop-filter: blur(5px);
        font-weight: 500;
    }

    /* ── Chat Bubble ── */
    .chat-bubble {
        background: rgba(255, 255, 255, 0.85);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        border-radius: 24px;
        padding: 2rem;
        box-shadow: 0 12px 35px rgba(30, 41, 59, 0.08);
        position: relative;
        margin-bottom: 2rem;
        border: 1px solid rgba(255,255,255,1);
        display: flex;
        align-items: center;
        gap: 1.8rem;
        transition: transform 0.3s;
    }
    .chat-bubble:hover {
        transform: translateY(-2px);
    }
    .chat-bubble h3 { color: #0F172A !important; margin: 0 0 0.5rem 0 !important; font-size: 1.6rem !important; font-weight: 800 !important; }
    .chat-bubble p  { color: #475569 !important; margin: 0 !important; font-size: 1.05rem !important; line-height: 1.6 !important; font-weight: 500; }

    /* ── Step Cards ── */
    .step-card {
        background: rgba(255, 255, 255, 0.7);
        backdrop-filter: blur(10px);
        border-radius: 20px;
        border: 1px solid rgba(255,255,255,0.8);
        padding: 1.5rem;
        margin-bottom: 1rem;
        box-shadow: 0 4px 15px rgba(30, 41, 59, 0.04);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }
    .step-card:hover { 
        box-shadow: 0 12px 30px rgba(30, 41, 59, 0.08); 
        transform: translateY(-3px);
        background: rgba(255, 255, 255, 0.9);
    }
    .step-title {
        font-size: 0.8rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        color: #334155;
        margin-bottom: 1.2rem;
        display: flex;
        align-items: center;
        gap: 10px;
    }
    .step-num {
        background: linear-gradient(135deg, #1E293B, #334155);
        color: #fff;
        width: 26px; height: 26px;
        border-radius: 50%;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 0.8rem;
        font-weight: 800;
        flex-shrink: 0;
        box-shadow: 0 4px 10px rgba(30,41,59,0.2);
    }

    /* ── Output Badge ── */
    .output-badge {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
        border-radius: 18px;
        padding: 1.2rem 1.5rem;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        gap: 1.2rem;
        color: #fff;
        box-shadow: 0 10px 30px rgba(15,23,42,0.3);
        transition: transform 0.3s;
    }
    .output-badge:hover { transform: translateY(-2px); }
    .output-badge h4 { margin: 0; font-size: 1.1rem; font-weight: 700; color: #fff; }
    .output-badge p  { margin: 0; font-size: 0.85rem; opacity: 0.8; color: #fff; }
    .output-stat {
        background: rgba(255,255,255,0.1);
        border-radius: 12px;
        padding: 10px 18px;
        font-size: 1.6rem;
        font-weight: 800;
        text-align: center;
        min-width: 75px;
        border: 1px solid rgba(255,255,255,0.1);
    }
    .output-stat span { display: block; font-size: 0.65rem; font-weight: 600; opacity: 0.8; margin-top: 2px; text-transform: uppercase; }

    /* ── Room Table Header ── */
    .room-header {
        display: grid;
        grid-template-columns: 1fr 100px 32px;
        gap: 8px;
        padding: 4px 4px 8px 4px;
        font-size: 0.7rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 1.2px;
        color: #64748B;
        border-bottom: 2px solid rgba(30,41,59,0.1);
        margin-bottom: 6px;
    }

    /* ── Summary Badges ── */
    .ct-row {
        display: flex;
        flex-wrap: wrap;
        gap: 10px;
        margin-bottom: 1rem;
    }
    .ct-badge {
        background: #F8FAFC;
        border: 1px solid #CBD5E1;
        border-radius: 20px;
        padding: 6px 16px;
        font-size: 0.8rem;
        font-weight: 700;
        color: #475569;
        box-shadow: 0 2px 5px rgba(0,0,0,0.02);
    }
    .ct-badge-active {
        background: #0F172A;
        border-color: #0F172A;
        color: #fff;
        box-shadow: 0 4px 10px rgba(15,23,42,0.2);
    }

    /* ── Inputs ── */
    .stTextInput > div > div > input {
        border-radius: 12px !important;
        border: 1px solid #CBD5E1 !important;
        padding: 0.7rem 1.2rem !important;
        background-color: #FFFFFF !important;
        color: #1E293B !important;
        font-weight: 500;
        transition: all 0.3s;
        box-shadow: 0 2px 5px rgba(0,0,0,0.02) !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: #6366F1 !important;
        box-shadow: 0 0 0 4px rgba(99,102,241,0.15) !important;
    }

    /* ── Checkbox ── */
    [data-testid="stCheckbox"] label span { color: #334155 !important; font-weight: 600 !important; }
    [data-testid="stCheckbox"] label { white-space: nowrap !important; }

    /* ── Main Button ── */
    div[data-testid="stButton"] > button {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%) !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 16px !important;
        padding: 0.9rem 2rem !important;
        font-weight: 800 !important;
        font-size: 1.1rem !important;
        width: 100%;
        box-shadow: 0 8px 25px rgba(15,23,42,0.3) !important;
        transition: all 0.3s cubic-bezier(0.34, 1.56, 0.64, 1) !important;
        letter-spacing: 0.5px;
    }
    div[data-testid="stButton"] > button:hover {
        transform: translateY(-3px) scale(1.02) !important;
        box-shadow: 0 15px 35px rgba(15,23,42,0.4) !important;
        background: linear-gradient(135deg, #1E293B 0%, #334155 100%) !important;
    }

    /* ── Download Button ── */
    div[data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #10B981 0%, #059669 100%) !important;
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 16px !important;
        padding: 0.9rem 2rem !important;
        font-weight: 800 !important;
        font-size: 1.1rem !important;
        width: 100%;
        box-shadow: 0 8px 25px rgba(16,185,129,0.3) !important;
        transition: all 0.3s cubic-bezier(0.34, 1.56, 0.64, 1) !important;
    }
    div[data-testid="stDownloadButton"] > button:hover {
        transform: translateY(-3px) scale(1.02) !important;
        box-shadow: 0 15px 35px rgba(16,185,129,0.4) !important;
    }

</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════
with st.sidebar:
    st.markdown(
        f"""<div class="mascot-container">{mascot_img_html}</div>
<h2 class="sidebar-title">MeowAI 🐾</h2>
<p class="sidebar-subtitle">Contract Extractor</p>""",
        unsafe_allow_html=True
    )

    st.markdown("""
<div class="sidebar-pill">
    📊 Output: <strong>43 Columns</strong> · Dashboard Ready
</div>
""", unsafe_allow_html=True)

    st.markdown("#### 🔑 Gemini API Key")
    
    import json
    import os
    
    KEYS_FILE = "user_keys.json"
    client_ip = "unknown"
    try:
        if hasattr(st, "context") and hasattr(st.context, "ip_address"):
            client_ip = st.context.ip_address or "unknown"
    except Exception:
        pass
        
    saved_key = ""
    try:
        if "GEMINI_API_KEY" in st.secrets:
            saved_key = st.secrets["GEMINI_API_KEY"]
    except Exception:
        pass

    if not saved_key and os.path.exists(KEYS_FILE):
        try:
            with open(KEYS_FILE, "r") as f:
                saved_key = json.load(f).get(client_ip, "")
        except Exception:
            pass

    api_key = st.text_input("API Key", value=saved_key, type="password", placeholder="AIza...", label_visibility="collapsed")
    
    if api_key and api_key != saved_key:
        try:
            saved_keys = {}
            if os.path.exists(KEYS_FILE):
                with open(KEYS_FILE, "r") as f:
                    saved_keys = json.load(f)
            saved_keys[client_ip] = api_key
            with open(KEYS_FILE, "w") as f:
                json.dump(saved_keys, f)
        except Exception:
            pass

    st.markdown("---")
    st.markdown("#### 🏨 Property Info")
    prop_type = st.radio("Property Type", ["Hotel / Resort", "Cruise Ship"])
    hotel_id  = st.text_input("Property ID", placeholder="e.g. 12711588")
    supplier  = st.text_input("Hotel Name", placeholder="e.g. InterContinental Khao Yai")

    st.markdown("---")
    st.markdown("""
<div style="text-align:center;color:#475569;font-size:0.72rem;padding:0.3rem 0;">
    🐾 Powered by Gemini 2.5 Flash<br>
    เหมียวดึงข้อมูลให้ครบ ไม่มีตกหล่น
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════
#  HERO
# ═══════════════════════════════════════════
st.markdown("""
<div class="chat-bubble">
    <div style="flex-shrink:0;font-size:3.2rem;animation:float 3s ease-in-out infinite;">😸</div>
    <div>
        <h3>สวัสดี DATA TEAM! 🐾</h3>
        <p>โยน PDF สัญญาโรงแรมหรือ Cruise มาเลย! หนู MeowAI จะสกัดข้อมูล Rate, Date, Surcharge, EB, Promo ทุก field ให้ครบ <strong>43 คอลัมน์</strong> ตามฟอร์แมต Dashboard พร้อม Upload ได้ทันที ไม่มีตกหล่น!</p>
    </div>
</div>
""", unsafe_allow_html=True)



# ═══════════════════════════════════════════
#  MAIN COLUMNS
# ═══════════════════════════════════════════
col_left, col_right = st.columns([1.1, 1], gap="large")

with col_left:
    # ── Step 1: Upload ──────────────────────
    st.markdown("""
<div class="step-card">
    <div class="step-title"><span class="step-num">1</span> อัพโหลดไฟล์สัญญา (PDF)</div>
</div>
""", unsafe_allow_html=True)
    uploaded_file = st.file_uploader(
        "ลากหรือเลือก PDF Contract",
        type=["pdf"],
        help="รองรับ PDF ภาษาอังกฤษ — Hotel, Resort และ Cruise"
    )

    # ── Step 2: Contract Types ──────────────
    st.markdown("""
<div class="step-card" style="margin-top:0.6rem;">
    <div class="step-title"><span class="step-num">2</span> ประเภทสัญญาที่จะสร้าง</div>
</div>
""", unsafe_allow_html=True)

    ct_main = st.checkbox("📋 Main Contract", value=True,
                          help="Base net rate — สร้างทุก period ใน PDF")

    ct_eb = st.checkbox("🦅 Early Bird", value=False,
                        help="AI จะ detect tier, days, discount% จาก PDF อัตโนมัติ — promo_code = E.B X DAYS, min_advance_days = X")

    c3, c4, c5 = st.columns([1.8, 1.1, 1.1])
    with c3:
        ct_promo = st.checkbox("🎯 Promotion", value=False,
                               help="Early Bird Offer หรือโปรโมชั่นพิเศษ")
    with c4:
        promo_code = st.text_input(
            "Promo Code", placeholder="EBO",
            disabled=not ct_promo, label_visibility="collapsed"
        )
    with c5:
        promo_till = st.text_input(
            "Book Till", placeholder="YYYY-MM-DD",
            disabled=not ct_promo, label_visibility="collapsed"
        )

    ct_por = st.checkbox("💎 POR (Price on Request)", value=False,
                         help="net_price=0, contract_type=POR")

    if ct_promo and promo_till:
        st.caption(f"📅 promo_book_till จะเป็น: `{promo_till} 23:59:59`")

with col_right:
    # ── Step 3: Rooms ───────────────────────
    st.markdown("""
<div class="step-card">
    <div class="step-title"><span class="step-num">3</span> ตั้งค่าห้องพัก / Cabin</div>
</div>
""", unsafe_allow_html=True)

    st.markdown("""
<div class="room-header">
    <span>Room Name</span>
    <span>Room ID</span>
    <span></span>
</div>
""", unsafe_allow_html=True)

    if "rooms" not in st.session_state:
        st.session_state.rooms = [{"room_name": "", "room_id": ""}]

    def add_room():
        st.session_state.rooms.append({"room_name": "", "room_id": ""})

    def remove_room(i):
        if len(st.session_state.rooms) > 1:
            st.session_state.rooms.pop(i)

    for idx, room in enumerate(list(st.session_state.rooms)):
        rc1, rc2, rc3 = st.columns([5, 3.5, 1])
        with rc1:
            room["room_name"] = st.text_input(
                f"room_name_{idx}", value=room["room_name"],
                key=f"rn_{idx}", placeholder="Deluxe Pool Villa",
                label_visibility="collapsed"
            )
        with rc2:
            room["room_id"] = st.text_input(
                f"room_id_{idx}", value=room["room_id"],
                key=f"ri_{idx}", placeholder="12711588001",
                label_visibility="collapsed"
            )
        with rc3:
            if st.button("✕", key=f"del_{idx}", help="ลบห้องนี้"):
                remove_room(idx)
                st.rerun()

    st.button("＋ เพิ่มห้องพัก", on_click=add_room, use_container_width=True)

    valid_count = sum(1 for r in st.session_state.rooms if r["room_name"].strip() and r["room_id"].strip())
    total_count = len(st.session_state.rooms)
    if valid_count:
        st.caption(f"✅ {valid_count}/{total_count} ห้องพร้อมใช้งาน")
    else:
        st.caption(f"⚠️ กรุณากรอก Room Name และ Room ID ให้ครบ")


# ═══════════════════════════════════════════
#  PROCESS LOGIC
# ═══════════════════════════════════════════
st.markdown("<br>", unsafe_allow_html=True)

valid_rooms = [r for r in st.session_state.rooms if r["room_name"].strip() and r["room_id"].strip()]

cts = []
if ct_main:  cts.append({"type": "main",  "label": "📋 Main Contract"})
if ct_eb:    cts.append({"type": "eb",    "label": "🦅 Early Bird"})
if ct_promo: cts.append({"type": "promo", "label": "🎯 Promotion",  "promo_code": promo_code, "promo_till": promo_till})
if ct_por:   cts.append({"type": "por",   "label": "💎 POR"})

# Summary badges
badge_html = ""
for c in cts:
    badge_html += f'<span class="ct-badge ct-badge-active">{c["label"]}</span>'
if valid_rooms:
    badge_html += f'<span class="ct-badge">🚪 {len(valid_rooms)} room(s)</span>'
if badge_html:
    st.markdown(f'<div class="ct-row">{badge_html}</div>', unsafe_allow_html=True)

ready = bool(api_key and uploaded_file and hotel_id and valid_rooms and cts)

if not ready:
    missing = []
    if not api_key:       missing.append("🔑 Gemini API Key")
    if not uploaded_file: missing.append("📄 ไฟล์ PDF")
    if not hotel_id:      missing.append("🏨 Property ID")
    if not valid_rooms:   missing.append("🚪 ห้องพัก (≥ 1 ห้อง)")
    if not cts:           missing.append("📋 ประเภทสัญญา")
    st.info("เมี้ยว~ ยังขาดข้อมูล: " + " · ".join(missing))

if st.button("🐾 สั่งเหมียวดึงข้อมูล — Generate Excel (43 columns)", use_container_width=True, key="process_btn"):
    if not ready:
        st.warning("⚠️ กรุณากรอกข้อมูลให้ครบก่อนนะเหมียว!")
    else:
        loading_overlay = st.empty()
        loading_overlay.markdown("""
        <style>
        .loading-overlay {
            position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
            background: rgba(15, 23, 42, 0.85); z-index: 999999;
            display: flex; flex-direction: column; justify-content: center; align-items: center;
            backdrop-filter: blur(8px);
        }
        .cat-spinner {
            font-size: 4rem;
            animation: bounce 1s infinite alternate;
        }
        .loading-bar-container {
            width: 300px; height: 10px; background: rgba(255,255,255,0.2);
            border-radius: 10px; margin-top: 20px; overflow: hidden;
        }
        .loading-bar {
            width: 50%; height: 100%; background: #6366F1;
            border-radius: 10px;
            animation: slide 1.5s infinite ease-in-out alternate;
        }
        .loading-text {
            color: #FFFFFF; font-size: 1.2rem; margin-top: 15px; font-weight: 600; letter-spacing: 1px;
            animation: pulse 1.5s infinite;
        }
        @keyframes bounce { from { transform: translateY(0); } to { transform: translateY(-20px); } }
        @keyframes slide { from { transform: translateX(-100%); } to { transform: translateX(200%); } }
        @keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.5; } }
        </style>
        <div class="loading-overlay">
            <div class="cat-spinner">🐱</div>
            <div class="loading-bar-container"><div class="loading-bar"></div></div>
            <div class="loading-text">กำลังแกะข้อมูล Contract... รอสักครู่นะเหมียว!</div>
        </div>
        """, unsafe_allow_html=True)
        try:
            file_bytes = uploaded_file.read()
            b64_data   = base64.b64encode(file_bytes).decode("utf-8")

            ct_desc_parts = []
            for c in cts:
                if c["type"] == "main":
                    ct_desc_parts.append("Main Contract: extract all standard net rates per period.")
                elif c["type"] == "eb":
                    ct_desc_parts.append(
                        'Early Bird: detect ALL tiers from PDF automatically. '
                        'For each tier, extract days_in_advance and discount_pct. '
                        'promo_code = "[Code from PDF] - E.B {days} DAYS" (e.g. "EBO26 - E.B 30 DAYS", or just "E.B 30 DAYS" if no code), '
                        'min_advance_days = days. Respect blackout dates (eb_blackout=true).'
                    )
                elif c["type"] == "promo":
                    ct_desc_parts.append(
                        f'Promotion: promo_code="{c["promo_code"]}", '
                        f'promo_book_till="{c["promo_till"]} 23:59:59".'
                    )
                elif c["type"] == "por":
                    ct_desc_parts.append('POR: net_price=0, promo_code="POR Rate".')
            ct_desc = " | ".join(ct_desc_parts)
    
            prompt_text = f"""You are an expert hotel contract data extractor. Extract ALL data from this PDF.
    
    ROOMS (use exactly these IDs and names, do not invent new ones): {json.dumps(valid_rooms)}
    CONTRACT TYPES TO EXTRACT: {ct_desc}
    PROPERTY TYPE: {prop_type}
    
    CRITICAL RULES:
    1. TRANSLATE EVERYTHING TO ENGLISH. All output must be in English, regardless of the original PDF language.
    2. Extract ONE period per distinct date range. Split periods if surcharge/blackout/min-nights rules differ.
    3. WEEKDAY/WEEKEND RATES:
       - Set has_weekday_weekend=true
       - rates dict = BASE (weekday) rate per room
       - has_surcharge=true
       - surcharge_rates = per-room weekend supplement dict e.g. {{"room_id_1": 1500, "room_id_2": 3000}}
       - If same surcharge for all rooms use surcharge_amount (single value) instead
       - Extract the exact weekend days stated in the PDF (e.g., "Friday & Saturday", "Saturday") into `weekend_days`. Extract the remaining weekday string into `weekday_days` (e.g., "Sun-Thu", "Sun-Fri").
       - If there is a surcharge (weekend, gala dinner, peak season, etc.), extract the exact reason/note into `surcharge_note` (e.g. "Room rate include surcharge on Christmas Eve").
    4. HTML FORMATTING PATTERNS (STRICT):
       Do NOT verbatim copy the PDF text. Summarize concisely, capture the exact meaning, and format STRICTLY using these HTML templates.
       - child_policy: (Do NOT include any food/meal-related information here. Split by room using `child_policies` dict)
     <p><span style="color: #008000;"><strong>Maximum Occupancy: [Occ]</strong></span></p>
     <p>Child [Age] years old Sharing bed + ABF = [Price/FOC] [Currency]</p>
     <p>Child/Adult Extra bed + ABF = [Price] [Currency]</p>
     <p><span style="color: #ff0000;"><strong>*Cannot add an extra bed</strong></span></p>
       - cancellation_policy: (Split by period in the periods list)
     <p><strong>Cancellation: [Season/Condition]</strong></p>
     <p>• Cancellation up to [X] days prior to arrival date, No charge.</p>
     <p><strong>No Show & Early Check-Out:</strong></p>
     <p>• The equivalent of the full originally booked length of stay will be charged.</p>
     (If strictly non-refundable, use: <p><span style="color:#f44336;"><strong>NON-REFUNDABLE</strong></span></p>)
       - meals_and_info:
     <p><strong>MAIN CONTRACT [YEAR] : [DATE] - [DATE]</strong> (Date format strictly e.g., 1 MAY 26 - 31 JUN 26)</p>
     <p><strong>※ MEAL PLAN</strong></p>
     <p>• [Details...]</p>
     <p><strong>※ MINIMUM NIGHTS & BLACKOUT DATES</strong></p>
     <p>• Minimum [X] Nights stay required on [DATES]</p>
     <p><span style="color: #008000;"><strong>COMPULSORY</strong></span> GALA DINNER [Details]</p>
     <p><strong>※ SUPPLEMENT CHARGE</strong> [Details]</p>
     <p><strong>※ EARLY BIRD & SPECIAL OFFERS</strong></p>
     <p>• [Details of Early bird or special offers... Check for specific advanced booking requirements and discount percentages.]</p>
     <p><span style="color: #ff0000;"><strong>Remark:</strong></span> Include any food space/location info here (e.g., at Somying's kitchen Restaurant).</p>
    5. promo_book_till format: "YYYY-MM-DD 23:59:59" (ONLY if PDF explicitly states a booking deadline)
    6. cutoff_date: Extract ONLY the raw number provided in the PDF (e.g., if "14 Days", output 14). Do NOT output a date.
    7. room_allotment: Extract as a dictionary mapping room_id to integer allotment (e.g., {{"room_id_1": 2, "room_id_2": 3}}).
    8. period_promo_note: Keep it concise. Do not add unnecessary info. Add ONLY critical period-specific conditions (e.g. MIN. 3 NIGHTS, Compulsory dinner).
    9. net_price, extra beds, and meals: output as integers ONLY (no decimals, no commas).
       - child_share_bed_abf, child_extra_bed_abf, extra_bed_abf, extra_bed_no_abf, full_board, half_board: Extract the prices as integer.
    10. rates dict key = room_id string EXACTLY as given in ROOMS above.
    11. PROMOTIONS: If the PDF explicitly gives promotion codes (e.g. "SUMMER26"), extract them into the `promotions` list. Calculate the discounted `rates` for each room and apply any extra bed/meal discounts if specified.
    12. CRUISE: detect night package from PDF (e.g. "1 Night", "2 Nights") and use as promo_code.
    13. hotel_id = "{hotel_id}", hotel_supplier = "{supplier}"
    
    Return ONLY valid JSON (no markdown fences, no explanation):
    {{
      "hotel_id": "{hotel_id}",
      "hotel_supplier": "{supplier}",
      "abf": "Included",
      "child_policy": null,
      "child_policies": {{"room_id_here": "<p>...</p>"}},
      "cancellation_policy": "<p>...</p>",
      "cutoff_date": 0,
      "meals_and_info": "<p><strong>MAIN CONTRACT...</strong></p>",
      "child_share_bed_abf": null,
      "child_extra_bed_abf": null,
      "extra_bed_abf": null,
      "extra_bed_no_abf": null,
      "full_board": null,
      "half_board": null,
      "rooms": {json.dumps(valid_rooms)},
      "promotions": [
    {{
      "promo_code": "PROMO26",
      "start_date": "YYYY-MM-DD",
      "end_date": "YYYY-MM-DD",
      "promo_note": null,
      "min_advance_days": null,
      "min_nights_stay": null,
      "rates": {{"room_id_here": 0}},
      "child_share_bed_abf": null,
      "child_extra_bed_abf": null,
      "extra_bed_abf": null,
      "extra_bed_no_abf": null,
      "full_board": null,
      "half_board": null
    }}
      ],
      "periods": [
    {{
      "start_date": "YYYY-MM-DD",
      "end_date": "YYYY-MM-DD",
      "season": "Season name",
      "is_peak": false,
      "eb_blackout": false,
      "has_surcharge": false,
      "surcharge_amount": 0,
      "surcharge_rates": {{}},
      "surcharge_currency": "THB",
      "surcharge_note": null,
      "period_promo_note": null,
      "cancellation_policy": "<p>...</p>",
      "min_nights_stay": null,
      "cutoff_date": 0,
      "room_allotment": {{"room_id_here": 2}},
      "has_weekday_weekend": false,
      "weekday_days": "Sun-Fri",
      "weekend_days": "Saturday",
      "rates": {{"room_id_here": 0}},
      "early_bird_tiers": [
        {{"days": 60, "discount_pct": 15, "promo_code": "E.B 60 DAYS", "promo_book_till": null}}
      ],
      "promotions": [
        {{"discount_pct": 20, "promo_code": "EBO", "promo_book_till": "YYYY-MM-DD 23:59:59"}}
      ]
    }}
      ]
    }}"""
    
            url = (
                "https://generativelanguage.googleapis.com/v1beta/"
                f"models/gemini-2.5-flash:generateContent?key={api_key}"
            )
            payload = {
                "contents": [{
                    "role": "user",
                    "parts": [
                        {"inlineData": {"mimeType": "application/pdf", "data": b64_data}},
                        {"text": prompt_text}
                    ]
                }],
                "generationConfig": {
                    "temperature": 0.1,
                    "responseMimeType": "application/json"
                }
            }
    
            import time
            max_retries = 5
            for attempt in range(max_retries):
                resp      = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=300)
                resp_data = resp.json()
    
                if resp.status_code != 200:
                    error_msg = resp_data.get("error", {}).get("message", "Gemini API Error")
                    if "high demand" in error_msg.lower() or resp.status_code >= 500:
                        if attempt < max_retries - 1:
                            # Exponential backoff: 5s, 10s, 20s, 40s
                            time.sleep(5 * (2 ** attempt))
                            continue
                    raise Exception(error_msg)
                else:
                    break
    
            text_output = resp_data["candidates"][0]["content"]["parts"][0]["text"]
            parsed_data = json.loads(text_output)
    
            # ── Generate rows (43-column dashboard format) ──
            rows = generate_rows(parsed_data, cts)
    
            # ── Build Excel workbook ──
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
    
            header_fill = PatternFill("solid", start_color="1F4E79")
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            data_font   = Font(name="Arial", size=10)
    
            # Header row — ALL 43 columns, always present
            for c_idx, h in enumerate(HEADERS, 1):
                cell = ws.cell(1, c_idx, h)
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(h, 12)
            ws.row_dimensions[1].height = 20
    
            date_only_cols = {"start_date", "end_date", "edited_date", "cutoff_date"}
            datetime_cols  = {"created_date"}
    
            # Data rows — ALL 43 columns, None → empty cell
            for r_idx, row in enumerate(rows, 2):
                for c_idx, h in enumerate(HEADERS, 1):
                    val  = row.get(h)
                    cell = ws.cell(r_idx, c_idx, val)
                    cell.font = data_font
                    if h in date_only_cols and val:
                        cell.number_format = "DD/MM/YYYY"
                    elif h in datetime_cols and val:
                        cell.number_format = "DD/MM/YYYY HH:MM:SS"
                    if h == "net_price" and val is not None:
                        cell.value = str(val)
    
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
    
            n_periods = len(parsed_data.get("periods", []))
            n_rooms   = len(valid_rooms)
            st.success(
                f"✅ สำเร็จ! เหมียวดึงข้อมูลได้ **{len(rows)} แถว** "
                f"({n_periods} periods × {n_rooms} rooms · 43 columns ครบทุก field)"
            )
            st.balloons()
            st.download_button(
                label="📥  Download Excel (.xlsx)",
                data=buf,
                file_name=f"Contract_{hotel_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
        except Exception as e:
            st.error(f"❌ เกิดข้อผิดพลาด: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            loading_overlay.empty()
